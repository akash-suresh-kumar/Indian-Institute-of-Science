import os
import time
import logging
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import win32com.client as win32
import pythoncom

# Set up logging
logging.basicConfig(filename='cost_cultivation_data_script.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def setup_chrome_driver(download_dir):
    options = Options()
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option('prefs', prefs)
    return webdriver.Chrome(options=options)

def navigate_to_year(driver, year):
    try:
        year_select = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, 'year_filter_sess'))
        )
        select = Select(year_select)
        target_year = f"{year}-{str(year+1)[-2:]}"
        
        if target_year not in [option.text for option in select.options]:
            archive_url = "https://desagri.gov.in/document-report-category/cost-of-cultivation-production-estimates-archive/"
            driver.get(archive_url)
            year_select = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, 'year_filter_sess'))
            )
            select = Select(year_select)
        
        if target_year not in [option.text for option in select.options]:
            logging.warning(f"Year {target_year} not found in both main and archive dropdowns. Skipping.")
            return False
        
        select.select_by_visible_text(target_year)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, '//table[@class="table"]'))
        )
        time.sleep(3)  # Reduced from 5 seconds to 3 seconds
        return True
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Error navigating to year {target_year}: {e}")
        return False

def download_excel_files(driver, download_dir):
    excel_links = driver.find_elements(By.XPATH, '//table[@class="table"]//a[contains(text(), "Download File (English)")]')
    if not excel_links:
        logging.warning("No 'Download File (English)' links found.")
        return

    for link in excel_links:
        try:
            link_url = link.get_attribute("href")
            if link_url:
                logging.info(f'Attempting to download: {link_url}')
                driver.execute_script("window.open('');")
                driver.switch_to.window(driver.window_handles[-1])
                driver.get(link_url)
                time.sleep(10)  # Reduced from 20 seconds to 10 seconds
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                logging.info(f'Initiated download for: {link_url}')
            else:
                logging.warning(f'Link URL is empty or missing')
        except Exception as e:
            logging.error(f'Error downloading link: {e}')

    time.sleep(15)  # Reduced from 30 seconds to 15 seconds

def process_excel_file(file_path):
    try:
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(file_path)
        wb.Activate()
        
        # Enable editing if read-only
        if excel.ActiveWorkbook.ReadOnly:
            excel.DisplayAlerts = False
            wb.SaveAs(file_path)
            excel.DisplayAlerts = True
        
        # Convert to .xlsx if it's .xls
        if file_path.endswith('.xls'):
            new_file_path = file_path + 'x'
            wb.SaveAs(new_file_path, FileFormat=51)  # 51 is for .xlsx
            wb.Close()
            excel.Quit()
            os.remove(file_path)  # Remove the original .xls file
            logging.info(f"Converted {file_path} to xlsx format")
            return new_file_path
        else:
            wb.Close()
            excel.Quit()
            return file_path
    except Exception as e:
        logging.error(f"Error processing file {file_path}: {str(e)}")
        return None
    finally:
        pythoncom.CoUninitialize()

def merge_excel_files(download_dir):
    excel_files = [f for f in os.listdir(download_dir) if f.endswith('.xlsx')]
    logging.info(f"Found {len(excel_files)} Excel files to process.")
    
    if not excel_files:
        logging.warning("No Excel files found in the download directory.")
        return None

    merged_workbook = Workbook()
    merged_workbook.remove(merged_workbook.active)  # Remove default sheet
    crop_data = {}

    for excel_file in excel_files:
        try:
            excel_path = os.path.join(download_dir, excel_file)
            logging.info(f"Processing file: {excel_file}")
            
            df = pd.read_excel(excel_path, sheet_name=None)
            year = excel_file.split('-')[0]  # Extract year from filename
            
            for sheet_name, sheet_df in df.items():
                sheet_df['Year'] = year
                sheet_df['Source_File'] = excel_file
                sheet_df['Original_Sheet'] = sheet_name
                
                crop_name = sheet_name.split('_')[0]  # Adjust based on actual naming convention
                
                if crop_name not in crop_data:
                    crop_data[crop_name] = []
                crop_data[crop_name].append(sheet_df)
            
            logging.info(f"Successfully processed data from {excel_file}")
        except Exception as e:
            logging.error(f"Error processing {excel_file}: {str(e)}")

    logging.info(f"Processed data for {len(crop_data)} crops.")

    for crop_name, data_list in crop_data.items():
        try:
            sheet = merged_workbook.create_sheet(title=crop_name[:31])  # Excel sheet name limit
            combined_df = pd.concat(data_list, ignore_index=True)
            for r in dataframe_to_rows(combined_df, index=False, header=True):
                sheet.append(r)
            logging.info(f"Added sheet for crop: {crop_name}")
        except Exception as e:
            logging.error(f"Error creating sheet for crop {crop_name}: {str(e)}")

    if len(merged_workbook.sheetnames) > 0:
        merged_file_path = os.path.join(download_dir, "merged_cost_cultivation_data.xlsx")
        try:
            merged_workbook.save(merged_file_path)
            logging.info(f"All Excel files merged into {merged_file_path}")
            return merged_file_path
        except Exception as e:
            logging.error(f"Error saving merged workbook: {str(e)}")
            return None
    else:
        logging.warning("No data was merged. The resulting workbook is empty.")
        return None

def convert_excel_to_csv(excel_file_path):
    try:
        # Create a directory for the CSV file
        csv_dir = os.path.dirname(excel_file_path)
        csv_file_path = os.path.join(csv_dir, "merged_cost_cultivation_data.csv")
        
        # Read all sheets from the Excel file
        xlsx = pd.ExcelFile(excel_file_path)
        all_sheets = []
        
        for sheet_name in xlsx.sheet_names:
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
            df['Sheet_Name'] = sheet_name  # Add a column to identify the original sheet
            all_sheets.append(df)
        
        # Concatenate all dataframes
        merged_df = pd.concat(all_sheets, ignore_index=True)
        
        # Save as CSV
        merged_df.to_csv(csv_file_path, index=False)
        logging.info(f"Converted merged Excel file to CSV: {csv_file_path}")
        
        return csv_file_path
    except Exception as e:
        logging.error(f"Error converting Excel to CSV: {str(e)}")
        return None

def process_cost_cultivation_data(start_year, end_year, download_dir):
    driver = setup_chrome_driver(download_dir)
    try:
        base_url = "https://desagri.gov.in/document-report-category/cost-of-cultivation-production-estimates/"
        for year in range(start_year, end_year + 1):
            logging.info(f"Processing year: {year}-{str(year+1)[-2:]}")
            driver.get(base_url)
            if navigate_to_year(driver, year):
                download_excel_files(driver, download_dir)
    except Exception as e:
        logging.error(f"An error occurred during download: {e}")
    finally:
        driver.quit()

    # Process and convert downloaded files
    for filename in os.listdir(download_dir):
        if filename.endswith(('.xls', '.xlsx')):
            file_path = os.path.join(download_dir, filename)
            process_excel_file(file_path)

    merged_file_path = merge_excel_files(download_dir)
    
    if merged_file_path:
        logging.info(f"Successfully created merged Excel file: {merged_file_path}")
        
        # Convert merged Excel file to CSV
        csv_file_path = convert_excel_to_csv(merged_file_path)
        if csv_file_path:
            logging.info(f"Successfully converted merged Excel file to CSV: {csv_file_path}")
            return merged_file_path, csv_file_path
        else:
            logging.error("Failed to convert merged Excel file to CSV.")
            return merged_file_path, None
    else:
        logging.error("Failed to create merged Excel file.")
    return None, None

def get_user_input(prompt, valid_responses):
    while True:
        response = input(prompt).lower().strip()
        if response in valid_responses:
            return response
        print(f"Please enter one of: {', '.join(valid_responses)}")

if __name__ == "__main__":
    download_dir = os.path.join(os.path.expanduser('~'), 'Downloads', 'cost_cultivation_data')
    
    print("Cost of Cultivation Data Processing Script")
    print("------------------------------------------")

    process_data = get_user_input("Do you want to download and merge Cost of Cultivation data? (y/n): ", ['y', 'n'])
    if process_data == 'y':
        start_year = int(input("Enter the start year (e.g., 2021 for 2021-22): "))
        end_year = int(input("Enter the end year (e.g., 2023 for 2023-24): "))
        print(f"Processing Cost of Cultivation data for years {start_year}-{str(start_year+1)[-2:]} to {end_year}-{str(end_year+1)[-2:]}...")
        merged_file_path, csv_file_path = process_cost_cultivation_data(start_year, end_year, download_dir)
        print("Processing completed.")
        
        if merged_file_path and os.path.exists(merged_file_path):
            print(f"Merged Excel file created: {merged_file_path}")
            if csv_file_path and os.path.exists(csv_file_path):
                print(f"CSV file created: {csv_file_path}")
            else:
                print("CSV file was not created. Please check the log for errors.")
        else:
            print("Merged Excel file was not created. Please check the log for errors.")

    print("Script execution finished.")
    print("Check the log file 'cost_cultivation_data_script.log' for detailed information.")